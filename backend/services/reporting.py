from __future__ import annotations

import os
from datetime import datetime
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook

DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "downloads")


def _ensure_download_dir() -> None:
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def _get_duration(series: Iterable[Dict[str, float]]) -> float:
    times = sorted(float(row.get("Elapsed Time", 0)) for row in series)
    if not times:
        return 0.0
    return float(times[-1] - times[0])


def _get_max_pressure(pressure_rows: Iterable[Dict[str, float]]) -> float:
    pressures = [float(row.get("Bladder Pressure", 0)) for row in pressure_rows]
    return float(max(pressures)) if pressures else 0.0


def _get_final_volume(volume_rows: Iterable[Dict[str, float]]) -> float:
    sorted_rows = sorted(volume_rows, key=lambda row: float(row.get("Elapsed Time", 0)))
    if not sorted_rows:
        return 0.0
    return float(sorted_rows[-1].get("Tot Infused Vol", 0))


def _append_kv_rows(ws, title: str, values: Optional[Dict[str, object]]) -> None:
    if not values:
        return
    ws.append([title, ""])  # spacer/title row
    for key, value in values.items():
        ws.append([key, value])


def create_report(
    data: Dict,
    peaks: Optional[List[Dict[str, float]]] = None,
    points: Optional[Dict[str, List[Dict[str, object]]]] = None,
    segments: Optional[List[Dict[str, object]]] = None,
    peak_params: Optional[Dict[str, object]] = None,
    segment_params: Optional[Dict[str, object]] = None,
    experiment_window: Optional[Dict[str, float]] = None,
) -> str:
    _ensure_download_dir()

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    duration = _get_duration(
        data.get("pressure") or data.get("scale") or data.get("volume") or []
    )
    max_pressure = _get_max_pressure(data.get("pressure", []))
    final_volume = _get_final_volume(data.get("volume", []))
    kept_intervals = data.get("kept_intervals")

    ws_summary.append(["Metric", "Value"])
    if experiment_window:
        ws_summary.append(["Experiment Window Start", experiment_window.get("start")])
        ws_summary.append(["Experiment Window End", experiment_window.get("end")])
    ws_summary.append(["Duration", duration])
    ws_summary.append(["Max Pressure", max_pressure])
    ws_summary.append(["Final Volume", final_volume])
    ws_summary.append(["Peak Count", len(peaks) if peaks else 0])
    ws_summary.append(["Segment Count", len(segments) if segments else 0])
    if kept_intervals is not None:
        ws_summary.append(["Number of kept intervals", kept_intervals])
    _append_kv_rows(ws_summary, "Peak Detection Params", peak_params)
    _append_kv_rows(ws_summary, "Onset/Empty Params", segment_params)
    ws_summary.append(["Generated", datetime.utcnow().isoformat()])

    ws_scale = wb.create_sheet("TimeSeries_Scale")
    ws_scale.append(["Elapsed Time", "Scale"])
    for row in data.get("scale", []):
        ws_scale.append([row.get("Elapsed Time"), row.get("Scale")])

    ws_volume = wb.create_sheet("TimeSeries_Volume")
    ws_volume.append(["Elapsed Time", "Tot Infused Vol"])
    for row in data.get("volume", []):
        ws_volume.append([row.get("Elapsed Time"), row.get("Tot Infused Vol")])

    ws_pressure = wb.create_sheet("TimeSeries_Pressure")
    ws_pressure.append(["Elapsed Time", "Bladder Pressure"])
    for row in data.get("pressure", []):
        ws_pressure.append([row.get("Elapsed Time"), row.get("Bladder Pressure")])

    ws_points = wb.create_sheet("Points")
    ws_points.append(["Type", "Time", "Value", "Index"])
    for peak in peaks or []:
        ws_points.append(["peak", peak.get("time"), peak.get("value"), peak.get("index")])
    for onset in (points or {}).get("onset", []):
        ws_points.append(["onset", onset.get("time"), onset.get("value"), onset.get("index")])
    for empty in (points or {}).get("empty", []):
        ws_points.append(["empty", empty.get("time"), empty.get("value"), empty.get("index")])

    ws_segments = wb.create_sheet("Segments")
    ws_segments.append(
        [
            "i",
            "onsetTime",
            "peakTime",
            "emptyTime",
            "imiSec",
            "maxPressure",
            "avgPressureBetweenEmptyAndNextOnset",
            "deltaVolume",
        ]
    )
    for segment in segments or []:
        metrics = segment.get("metrics") or {}
        ws_segments.append(
            [
                segment.get("i"),
                segment.get("onsetTime"),
                segment.get("peakTime"),
                segment.get("emptyTime"),
                metrics.get("imiSec"),
                metrics.get("maxPressure"),
                metrics.get("avgPressureBetweenEmptyAndNextOnset"),
                metrics.get("deltaVolume"),
            ]
        )

    filename = f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(DOWNLOAD_DIR, filename)
    wb.save(filepath)

    return filename
